"""
双向同步脚本：
  python sync_stats.py          # 代码 -> 表格（导出）
  python sync_stats.py export   # 同上
  python sync_stats.py import   # 表格 -> 代码（导入）
"""
import re, sys
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

MODE = sys.argv[1] if len(sys.argv) > 1 else 'export'
HTML_FILE = 'index.html'
XLSX_FILE = 'game_stats.xlsx'

# ========== 公共工具 ==========
hf = Font(bold=True, size=11, color='FFFFFF')
border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

def write_header(ws, headers, fill_color='1e293b'):
    fill = PatternFill('solid', fgColor=fill_color)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hf; cell.fill = fill
        cell.alignment = Alignment(horizontal='center'); cell.border = border

def write_row(ws, row_idx, data):
    for c, v in enumerate(data, 1):
        cell = ws.cell(row=row_idx, column=c, value=v)
        cell.border = border; cell.alignment = Alignment(horizontal='center', wrap_text=True)

# ========== EXPORT: 代码 -> 表格 ==========
def do_export():
    html = open(HTML_FILE, encoding='utf-8').read()

    # 解析 ENEMY_TYPES
    m = re.search(r'const ENEMY_TYPES\s*=\s*\{(.+?)\};', html, re.S)
    types_block = m.group(1) if m else ''
    enemy_data = {}
    for em in re.finditer(r"(\w+)\s*:\s*\{(.+?)\}", types_block):
        key = em.group(1)
        props = {}
        for p in re.finditer(r"(\w+)\s*:\s*([\w.]+|true|false)", em.group(2)):
            v = p.group(2)
            if v == 'true': v = True
            elif v == 'false': v = False
            else:
                try: v = float(v) if '.' in v else int(v, 0 if v.startswith('0x') else 10)
                except: pass
            props[p.group(1)] = v
        enemy_data[key] = props

    # 解析 CONFIG
    cfg = {}
    cfg_block = html[html.find('const CONFIG'):html.find('const CONFIG')+1200]
    for cm in re.finditer(r"(\w+)\s*:\s*([\d.]+)", cfg_block):
        try: cfg[cm.group(1)] = float(cm.group(2)) if '.' in cm.group(2) else int(cm.group(2))
        except: pass

    # 解析 buildWaveQueue
    bwq = re.search(r'const baseCount\s*=\s*(\d+)', html)
    baseCount = int(bwq.group(1)) if bwq else 30
    gwf = re.search(r'const growthFactor\s*=\s*([\d.]+)', html)
    growthFactor = float(gwf.group(1)) if gwf else 1.22

    # 解析 player
    player = {}
    p_block = html[html.find('const player'):html.find('const player')+500]
    for pm in re.finditer(r"(\w+)\s*:\s*([\d.]+)", p_block):
        try: player[pm.group(1)] = float(pm.group(2)) if '.' in pm.group(2) else int(pm.group(2))
        except: pass

    # Boss HP
    bhp = re.search(r'const bossHp\s*=\s*(\d+)\s*\+\s*gameState\.level\s*\*\s*(\d+)', html)
    bossHpBase = int(bhp.group(1)) if bhp else 500
    bossHpPerLv = int(bhp.group(2)) if bhp else 50

    # 经验系数
    exp_c = re.search(r'e\.userData\.score\s*\*\s*([\d.]+)\s*\*', html)
    expCoeff = float(exp_c.group(1)) if exp_c else 0.3

    # ===== 生成 Excel =====
    wb = Workbook()

    # Sheet 1: 怪物数值（可编辑列：体型/HP/移速/金币/分数）
    ws1 = wb.active; ws1.title = '怪物数值'
    write_header(ws1, ['类型ID','中文名','体型','基础HP','移速','金币','分数','飞行','视觉特征'])
    names = {'grunt':'步兵','fast':'突击兵','flyer':'飞行兵','tank':'重装兵','boss':'BOSS'}
    visuals = {'grunt':'头顶红色小角','fast':'尖头+粉色拖尾','flyer':'紫色翅膀+推进器','tank':'蓝色护盾+肩甲','boss':'双角+发光眼+光环'}
    yellow = PatternFill('solid', fgColor='FFF3CD')
    for i, (k, v) in enumerate(enemy_data.items(), 2):
        hp_val = f"{bossHpBase}+Lv*{bossHpPerLv}" if k == 'boss' else v.get('hp', 1)
        write_row(ws1, i, [k, names.get(k,k), v.get('size'), hp_val, v.get('speed'),
            v.get('gold'), v.get('score'), '是' if v.get('fly') else '否', visuals.get(k,'')])
        # 标记可编辑列为黄色底
        for col in [3,4,5,6,7]:
            ws1.cell(row=i, column=col).fill = yellow
    # 说明行
    ws1.cell(row=len(enemy_data)+3, column=1, value='黄色底 = 可编辑，import 时会同步回代码').font = Font(italic=True, color='888888')
    for w,width in [('A',8),('B',10),('C',8),('D',14),('E',8),('F',8),('G',8),('H',6),('I',28)]:
        ws1.column_dimensions[w].width = width

    # Sheet 2: 波次规则（可编辑）
    ws2 = wb.create_sheet('波次规则')
    write_header(ws2, ['参数','值','说明'])
    wave_data = [
        ['baseCount', baseCount, '基础出怪数'],
        ['growthFactor', growthFactor, '每波增长因子'],
        ['restEveryNWaves', cfg.get('restEveryNWaves',5), '每N波休整'],
        ['restDuration', cfg.get('restDuration',20), '休整时长(秒)'],
        ['bossWave', cfg.get('bossWave',25), 'Boss波次'],
        ['bossUnlockWave', cfg.get('bossUnlockWave',15), 'Boss门解锁波次'],
        ['maxRevives', cfg.get('maxRevives',5), '复活次数(阶段一)'],
        ['expBase', cfg.get('expBase',50), '升2级所需经验'],
        ['expGrowth', cfg.get('expGrowth',1.35), '经验增长率'],
        ['expCoeff', expCoeff, '击杀经验系数'],
        ['bossHpBase', bossHpBase, 'Boss基础HP'],
        ['bossHpPerLv', bossHpPerLv, 'Boss每级+HP'],
    ]
    for i, r in enumerate(wave_data, 2):
        write_row(ws2, i, r)
        ws2.cell(row=i, column=2).fill = yellow
    ws2.column_dimensions['A'].width = 18; ws2.column_dimensions['B'].width = 14; ws2.column_dimensions['C'].width = 25

    # Sheet 3: Boss技能
    ws3 = wb.create_sheet('Boss技能')
    write_header(ws3, ['技能','触发','效果','伤害','备注'], 'DC2626')
    for i, r in enumerate([
        ['追踪','默认','追踪玩家','-','移速2.5'],
        ['冲撞','10s循环','冲刺1.5s','8','碰墙停'],
        ['召唤尸潮','10s循环','召唤6只小怪','-',''],
        ['AOE践踏','10s循环','12m范围','最高25',''],
        ['护盾','HP<50%','5s减伤50%','-','30sCD'],
    ], 2): write_row(ws3, i, r)
    ws3.column_dimensions['A'].width=10; ws3.column_dimensions['B'].width=10
    ws3.column_dimensions['C'].width=20; ws3.column_dimensions['D'].width=8; ws3.column_dimensions['E'].width=10

    # Sheet 4: 玩家数值
    ws4 = wb.create_sheet('玩家数值')
    write_header(ws4, ['参数','值','说明'], '0891B2')
    for i, r in enumerate([
        ['speed', player.get('speed',7.5), '移动速度'],
        ['jumpSpeed', player.get('jumpSpeed',6.0), '跳跃初速'],
        ['gravity', player.get('gravity',20), '重力'],
    ], 2):
        write_row(ws4, i, r)
        ws4.cell(row=i, column=2).fill = yellow
    ws4.column_dimensions['A'].width=14; ws4.column_dimensions['B'].width=10; ws4.column_dimensions['C'].width=16

    wb.save(XLSX_FILE)
    print(f'[export] 已从代码同步到 {XLSX_FILE}')

# ========== IMPORT: 表格 -> 代码 ==========
def do_import():
    wb = load_workbook(XLSX_FILE, data_only=True)
    html = open(HTML_FILE, encoding='utf-8').read()
    changed = []

    # Sheet 1: 怪物数值 -> ENEMY_TYPES
    ws1 = wb['怪物数值']
    for row in ws1.iter_rows(min_row=2, max_row=6, values_only=False):
        vals = [c.value for c in row]
        tid = vals[0]  # 类型ID
        if not tid or tid not in ['grunt','fast','flyer','tank','boss']: continue
        size, hp, speed, gold, score = vals[2], vals[3], vals[4], vals[5], vals[6]
        if tid == 'boss': continue  # Boss HP 特殊处理
        # 构建替换正则
        pat = re.compile(
            rf"({tid}\s*:\s*\{{[^}}]*?size\s*:\s*)[\d.]+([^}}]*?hp\s*:\s*)[\d.]+([^}}]*?speed\s*:\s*)[\d.]+([^}}]*?gold\s*:\s*)[\d.]+([^}}]*?score\s*:\s*)[\d.]+",
            re.S
        )
        def repl(m):
            return f"{m.group(1)}{size}{m.group(2)}{hp}{m.group(3)}{speed}{m.group(4)}{gold}{m.group(5)}{score}"
        new_html = pat.sub(repl, html)
        if new_html != html:
            changed.append(f"  {tid}: size={size} hp={hp} speed={speed} gold={gold} score={score}")
            html = new_html

    # Sheet 2: 波次规则 -> CONFIG + buildWaveQueue + Boss HP + 经验系数
    ws2 = wb['波次规则']
    params = {}
    for row in ws2.iter_rows(min_row=2, max_row=20, values_only=True):
        if row[0] and row[1] is not None:
            params[row[0]] = row[1]

    cfg_keys = ['restEveryNWaves','restDuration','bossWave','bossUnlockWave','maxRevives','expBase','expGrowth']
    for key in cfg_keys:
        if key in params:
            val = params[key]
            old = re.search(rf"({key}\s*:\s*)([\d.]+)", html)
            if old and str(old.group(2)) != str(val):
                html = re.sub(rf"({key}\s*:\s*)[\d.]+", rf"\g<1>{val}", html, count=1)
                changed.append(f"  CONFIG.{key}: {old.group(2)} -> {val}")

    if 'baseCount' in params:
        html = re.sub(r'(const baseCount\s*=\s*)\d+', rf"\g<1>{int(params['baseCount'])}", html)
        changed.append(f"  baseCount -> {int(params['baseCount'])}")
    if 'growthFactor' in params:
        html = re.sub(r'(const growthFactor\s*=\s*)[\d.]+', rf"\g<1>{params['growthFactor']}", html)
        changed.append(f"  growthFactor -> {params['growthFactor']}")
    if 'expCoeff' in params:
        html = re.sub(r'(e\.userData\.score\s*\*\s*)[\d.]+(\s*\*)', rf"\g<1>{params['expCoeff']}\2", html)
        changed.append(f"  expCoeff -> {params['expCoeff']}")
    if 'bossHpBase' in params:
        html = re.sub(r'(const bossHp\s*=\s*)\d+', rf"\g<1>{int(params['bossHpBase'])}", html)
        changed.append(f"  bossHpBase -> {int(params['bossHpBase'])}")
    if 'bossHpPerLv' in params:
        html = re.sub(r'(gameState\.level\s*\*\s*)\d+', rf"\g<1>{int(params['bossHpPerLv'])}", html)
        changed.append(f"  bossHpPerLv -> {int(params['bossHpPerLv'])}")

    # Sheet 4: 玩家数值 -> player
    if '玩家数值' in wb.sheetnames:
        ws4 = wb['玩家数值']
        for row in ws4.iter_rows(min_row=2, max_row=10, values_only=True):
            if row[0] and row[1] is not None:
                key, val = row[0], row[1]
                if key in ['speed','jumpSpeed','gravity']:
                    old = re.search(rf"({key}\s*:\s*)([\d.]+)", html[html.find('const player'):html.find('const player')+500])
                    if old:
                        html = html[:html.find('const player')] + re.sub(rf"({key}\s*:\s*)[\d.]+", rf"\g<1>{val}", html[html.find('const player'):html.find('const player')+500]) + html[html.find('const player')+500:]
                        changed.append(f"  player.{key}: {old.group(2)} -> {val}")

    if changed:
        open(HTML_FILE, 'w', encoding='utf-8').write(html)
        print(f'[import] 已同步 {len(changed)} 项到代码:')
        for c in changed: print(c)
    else:
        print('[import] 无变化')

if MODE == 'import':
    do_import()
else:
    do_export()
